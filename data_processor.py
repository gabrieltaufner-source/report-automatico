"""
Reads client data from Google Sheets (preferred) or local xlsx fallback.
"""
from datetime import datetime, date
from typing import Optional, Tuple
import re
import openpyxl


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_date(value) -> Optional[date]:
    """Return a date if value looks like dd/mm/yyyy or is a datetime, else None."""
    try:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        s = str(value).strip()
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
        if m:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except (ValueError, TypeError):
        pass
    return None


def _parse_period(period_str: str, ref_year: Optional[int] = None) -> Tuple[date, date]:
    """
    Parse '14/04 a 20/04' or '14/04/2025 a 20/04/2025' → (start, end).
    Uses ref_year (or current year) when year is omitted.
    """
    year = ref_year or datetime.now().year
    parts = re.split(r"\s+a\s+", period_str.strip(), flags=re.IGNORECASE)
    if len(parts) != 2:
        raise ValueError(f"Formato de período inválido: {period_str!r}. Use 'DD/MM a DD/MM'.")

    def parse_part(p):
        p = p.strip()
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", p)
        if not m:
            raise ValueError(f"Data inválida: {p!r}")
        d, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3) or year)
        return date(yr, mo, d)

    start, end = parse_part(parts[0]), parse_part(parts[1])
    return start, end


def _get_worksheet(wb):
    if "Acompanhamento Geral" in wb.sheetnames:
        return wb["Acompanhamento Geral"]
    return wb.active


def _find_header_in_raw(raw_rows: list) -> Tuple[int, dict]:
    """Find header row index and col_map in a list-of-lists (from Sheets API)."""
    for i, row in enumerate(raw_rows):
        for j, cell in enumerate(row):
            if str(cell).strip().upper() == "DATA":
                col_map = {str(v).strip().upper(): idx for idx, v in enumerate(row) if v}
                return i, col_map
    raise ValueError("Coluna 'DATA' não encontrada.")


def _find_header_row(ws):
    """Return (row_index, col_map) where row_index is 1-based header row."""
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().upper() == "DATA":
                col_map = {}
                for c in ws.iter_rows(min_row=cell.row, max_row=cell.row):
                    for cc in c:
                        if cc.value is not None:
                            col_map[str(cc.value).strip().upper()] = cc.column - 1
                return cell.row, col_map
    raise ValueError("Coluna 'DATA' não encontrada na planilha.")


def _rows_from_raw(raw_rows: list, header_idx: int, col_map: dict) -> list:
    """Convert Sheets API list-of-lists into row dicts, skipping non-date rows."""
    rows = []
    for row in raw_rows[header_idx + 1:]:
        if not row:
            continue
        raw_date = row[col_map["DATA"]] if col_map["DATA"] < len(row) else None
        if not raw_date:
            continue
        s = str(raw_date).strip()
        if s.upper() == s and not re.search(r"\d", s):
            continue
        if s.upper() == "TOTAL":
            continue
        d = _parse_date(raw_date)
        if d is None:
            continue
        row_dict = {"_date": d}
        for col_name, col_idx in col_map.items():
            row_dict[col_name] = row[col_idx] if col_idx < len(row) else None
        rows.append(row_dict)
    return rows


def _read_rows(ws, header_row: int, col_map: dict) -> list:
    """Return all data rows as dicts, skipping non-date rows."""
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        raw_date = row[col_map["DATA"]] if "DATA" in col_map else None
        if raw_date is None:
            continue
        s = str(raw_date).strip()
        if s.upper() == s and not re.search(r"\d", s):
            continue
        if s.upper() == "TOTAL":
            continue
        d = _parse_date(raw_date)
        if d is None:
            continue
        row_dict = {"_date": d}
        for col_name, col_idx in col_map.items():
            if col_idx < len(row):
                row_dict[col_name] = row[col_idx]
        rows.append(row_dict)
    return rows


def _load_all_rows(xlsx_path: str, sheet_id: Optional[str]) -> list:
    """Load rows from Google Sheets if sheet_id provided, else from local xlsx."""
    if sheet_id:
        from google_sheets import read_sheet
        raw = read_sheet(sheet_id)
        header_idx, col_map = _find_header_in_raw(raw)
        return _rows_from_raw(raw, header_idx, col_map)
    else:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = _get_worksheet(wb)
        header_row, col_map = _find_header_row(ws)
        return _read_rows(ws, header_row, col_map)


def _filter_rows(rows: list[dict], start: date, end: date) -> list[dict]:
    return [r for r in rows if start <= r["_date"] <= end]


def _safe_float(v) -> float:
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace("R$", "").replace("\xa0", "").replace(" ", "").replace("%", "").strip()
        if not s or s == "-" or s == "—":
            return 0.0
        if "," in s and "." in s:
            # Formato brasileiro: 1.234,56 → remove . (milhar) e troca , por .
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            # Só vírgula: 1234,56 → 1234.56
            s = s.replace(",", ".")
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def _fmt_money(v: float) -> str:
    return "R$ {:,.2f}".format(v).replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_pct(v: float) -> str:
    return f"{v:.2f}%".replace(".", ",")


def _fmt_count(v: float) -> str:
    return str(int(round(v)))


def _fmt_comp(current: float, previous: float, fmt_fn) -> str:
    """Format a comparison value with ▲/▼ and percentage change."""
    if previous == 0:
        return "—"
    pct = (current - previous) / abs(previous) * 100
    arrow = "▲" if pct >= 0 else "▼"
    sign = "+" if pct >= 0 else ""
    return f"{arrow} {sign}{pct:.1f}%".replace(".", ",")


# ── Ecommerce ─────────────────────────────────────────────────────────────────

def process_ecommerce(xlsx_path: str, periodo: str, periodo_comp: str, metas: dict, sheet_id: Optional[str] = None) -> dict:
    all_rows = _load_all_rows(xlsx_path, sheet_id)

    start, end = _parse_period(periodo)
    c_start, c_end = _parse_period(periodo_comp)

    cur = _filter_rows(all_rows, start, end)
    prv = _filter_rows(all_rows, c_start, c_end)

    # Monthly: all rows in same month as 'end'
    month_rows = [r for r in all_rows if r["_date"].year == end.year and r["_date"].month == end.month and r["_date"] <= end]

    def agg(rows):
        inv = sum(_safe_float(r.get("VALOR INVESTIDO", 0)) for r in rows)
        fat = sum(_safe_float(r.get("FATURAMENTO", 0)) for r in rows)
        ped = sum(_safe_float(r.get("PEDIDOS", 0)) for r in rows)
        ses = sum(_safe_float(r.get("SESSÕES", 0)) for r in rows)
        roas = fat / inv if inv else 0
        cps = inv / ped if ped else 0
        cpa = cps  # CPA = CPS (custo por aquisição = custo por pedido)
        taxa = ped / ses * 100 if ses else 0
        tck = fat / ped if ped else 0
        return dict(inv=inv, fat=fat, ped=ped, ses=ses, roas=roas, cps=cps, cpa=cpa, taxa=taxa, tck=tck)

    c = agg(cur)
    p = agg(prv)
    m = agg(month_rows)
    inv_mes = sum(_safe_float(r.get("VALOR INVESTIDO", 0)) for r in month_rows)
    fat_mes = sum(_safe_float(r.get("FATURAMENTO", 0)) for r in month_rows)

    meta_fat = float(metas.get("faturamento_mensal", 0))
    meta_inv = float(metas.get("investimento_mensal", 0))
    per_meta_fat = fat_mes / meta_fat * 100 if meta_fat else 0
    per_meta_inv = inv_mes / meta_inv * 100 if meta_inv else 0

    return {
        "fat_sem": _fmt_money(c["fat"]),
        "fat_sem_comp": _fmt_comp(c["fat"], p["fat"], _fmt_money),
        "inv_sem": _fmt_money(c["inv"]),
        "inv_sem_comp": _fmt_comp(c["inv"], p["inv"], _fmt_money),
        "roas": f"{c['roas']:.2f}".replace(".", ","),
        "roas_comp": _fmt_comp(c["roas"], p["roas"], str),
        "taxa_conv": _fmt_pct(c["taxa"]),
        "taxa_conv_comp": _fmt_comp(c["taxa"], p["taxa"], str),
        "vendas": _fmt_count(c["ped"]),
        "vendas_comp": _fmt_comp(c["ped"], p["ped"], str),
        "cps": _fmt_money(c["cps"]),
        "cps_comp": _fmt_comp(c["cps"], p["cps"], str),
        "tck_med": _fmt_money(c["tck"]),
        "tck_med_comp": _fmt_comp(c["tck"], p["tck"], str),
        "cpa": _fmt_money(c["cpa"]),
        "cpa_comp": _fmt_comp(c["cpa"], p["cpa"], str),
        "fat_mes": _fmt_money(fat_mes),
        "meta_fat": _fmt_money(meta_fat),
        "per_meta_fat": _fmt_pct(per_meta_fat),
        "inv_mes": _fmt_money(inv_mes),
        "meta_inv": _fmt_money(meta_inv),
        "per_meta_inv": _fmt_pct(per_meta_inv),
    }


# ── Lead ──────────────────────────────────────────────────────────────────────

def process_lead(xlsx_path: str, periodo: str, periodo_comp: str, metas: dict, sheet_id: Optional[str] = None) -> dict:
    all_rows = _load_all_rows(xlsx_path, sheet_id)

    start, end = _parse_period(periodo)
    c_start, c_end = _parse_period(periodo_comp)

    cur = _filter_rows(all_rows, start, end)
    prv = _filter_rows(all_rows, c_start, c_end)

    month_rows = [r for r in all_rows if r["_date"].year == end.year and r["_date"].month == end.month and r["_date"] <= end]

    def agg(rows):
        inv = sum(_safe_float(r.get("VALOR INVESTIDO", 0)) for r in rows)
        leads = sum(_safe_float(r.get("LEADS", 0)) for r in rows)
        cpl = inv / leads if leads else 0
        return dict(inv=inv, leads=leads, cpl=cpl)

    c = agg(cur)
    p = agg(prv)

    leads_mes = sum(_safe_float(r.get("LEADS", 0)) for r in month_rows)
    inv_mes = sum(_safe_float(r.get("VALOR INVESTIDO", 0)) for r in month_rows)

    meta_leads = float(metas.get("leads_mensal", 0))
    meta_inv = float(metas.get("investimento_mensal", 0))
    per_meta_leads = leads_mes / meta_leads * 100 if meta_leads else 0
    per_meta_inv = inv_mes / meta_inv * 100 if meta_inv else 0

    return {
        "lead_sem": _fmt_count(c["leads"]),
        "lead_sem_comp": _fmt_comp(c["leads"], p["leads"], str),
        "inv_sem": _fmt_money(c["inv"]),
        "inv_sem_comp": _fmt_comp(c["inv"], p["inv"], _fmt_money),
        "cpl": _fmt_money(c["cpl"]),
        "cpl_comp": _fmt_comp(c["cpl"], p["cpl"], _fmt_money),
        "leads_mes": _fmt_count(leads_mes),
        "meta_leads": _fmt_count(meta_leads),
        "per_meta_leads": _fmt_pct(per_meta_leads),
        "inv_mes": _fmt_money(inv_mes),
        "meta_inv": _fmt_money(meta_inv),
        "per_meta_inv": _fmt_pct(per_meta_inv),
    }
